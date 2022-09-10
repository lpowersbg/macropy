
from openpyxl import Workbook, load_workbook

hLink = load_workbook('HomeOffense.xlsx')
aLink = load_workbook('AwayOffense.xlsx')

#stat tracker
#wb.save(sheet)

def fillSheetPoints(player, points, sheet):
    #print(player)
    #print(points)
    
    for i in range(1, 70):
        aRow = 'A' + str(i)
        inputRow = 'G' + str(i)
        fName = 'D' + str(i)
        lName = 'E' + str(i)

        #print(aRow)
        #print(inputRow)
        
        if str(sheet[aRow].value) == player:
            if sheet[inputRow].value == None:
                sheet[inputRow].value = 0

            #print(sheet[inputRow].value)
            
            sheet[inputRow].value = sheet[inputRow].value + int(points)

            print("Total points for " + str(sheet[fName].value) + " " + str(sheet[lName].value) + " is " + str(sheet[inputRow].value))


def getSheet(nameGiven, ws):
    for i in ws:
        print(i)
        if nameGiven == i.title:
            return i
        
    
def start():

    home = hLink.active
    away = aLink.active


    team = input("Team: ")
    team == str.lower(team)
    
    player = input("Player: ")

    points = input("Points: ")

    
    if team == "h":
        #print("Home team")
        print("HOME player #" + player + " scored " + points + "." )
        fillSheetPoints(player, points, home)
    elif team == "a":
        #print("Away team")
        print("AWAY player #" + player + " scored " + points + "." )
        fillSheetPoints(player, points, away)

    hLink.save('HomeOffense.xlsx')
    aLink.save('AwayOffense.xlsx')
    

    ##cont = input("Continue? Y/N: ")
    print(" ")
    print(" ")
    ##cont = str.lower(cont)

   ## if cont == "y":
     ##   start()

    start()


if __name__ == "__main__":
    start()
    

